import openpyxl
wb = openpyxl.load_workbook("C:\Selenium_Python\SelPackage\ReadData.xlsx")
sh = wb["Login"]
rows = sh.max_row
cols = sh.max_column
print(rows,cols)
for r in range(1, rows+1):
    for col in range(1,cols+1):
        val = sh.cell(row=r,column=col).value
        print(val,end="  ")
    print()